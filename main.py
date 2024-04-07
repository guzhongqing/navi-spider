import os
import time

import requests
from bs4 import BeautifulSoup

# Excel写入函数
from openpyxl import Workbook, load_workbook


def write_excel(title, title_url, category, publish_time, view, comment, like):
    filename = "blog.xlsx"

    # 检查文件是否存在
    if os.path.exists(filename):
        # 加载现有工作簿
        workbook = load_workbook(filename)
        worksheet = workbook.active
        # 新数据写入下一行
        row = worksheet.max_row + 1
    else:
        # 创建新工作簿和工作表
        workbook = Workbook()
        worksheet = workbook.active
        # 设置列标题
        headers = ["文章标题", "文章地址", "文章分类", "发布时间", "阅读量", "评论量", "点赞量"]
        worksheet.append(headers)
        row = 2  # 从第二行开始写入数据

    # 写入数据
    data = [title, title_url, category, publish_time, view, comment, like]
    worksheet.append(data)

    # 设置URL单元格的超链接和样式
    url_cell = worksheet.cell(row=row, column=2)
    url_cell.hyperlink = title_url
    url_cell.value = title_url
    url_cell.style = "Hyperlink"  # 使用内置的"Hyperlink"样式

    # 保存工作簿
    workbook.save(filename)
    print("写入成功")


# 日期转换函数 "2024年4月6日 12:17发布"
def date_convert(date):
    from datetime import datetime

    # 移除"发布"并替换中文日期部分，以匹配datetime的格式要求
    date_str = (
        date.replace("发布", "").replace("年", "-").replace("月", "-").replace("日", "")
    )

    # %Y, %m, %d, %H, %M 分别代表年、月、日、小时和分钟
    date_obj = datetime.strptime(date_str, "%Y-%m-%d %H:%M")
    return date_obj


# 输入字符串可能包含数字、"K"（代表千）或"M"（代表百万）
# 需要根据后缀来决定如何转换
def convert_str_to_num(input_str):
    # 去除空格
    cleaned_str = input_str.strip()

    # 检查是否以"K"或"M"结尾并相应地替换
    if cleaned_str.endswith("K") or cleaned_str.endswith("k"):
        return float(cleaned_str[:-1]) * 1e3
    elif cleaned_str.endswith("M") or cleaned_str.endswith("m"):
        return float(cleaned_str[:-1]) * 1e6
    else:
        return float(cleaned_str)


# 根据地址获取当页面博客列表数据
def get_blog_list(url):
    # 发送请求到博客页面
    response: requests.Response = requests.get(url)

    # print(response.text)
    soup = BeautifulSoup(response.text, "html.parser")

    # 使用CSS选择器获取元素
    elements = soup.select(
        "#content > main > div > div > div.cat_list > div.list-grid.list-grid-padding"
    )

    # 遍历找到的元素并打印或处理
    for element in elements:
        # 获取博客列表地址
        blog_a = element.select_one("h2>a")
        title_url = blog_a["href"]
        print(title_url)

        # 根据地址进入文章
        response: requests.Response = requests.get(blog_a["href"])
        soup = BeautifulSoup(response.text, "html.parser")

        main_content_element = soup.select_one(
            "#content > main > div.content-wrap > div > div.panel.card > div > div.panel-header.mb-4"
        )
        # print(main_content_element.text)
        # 获取文章相关数据
        title = main_content_element.select_one("h1").text
        category = main_content_element.select_one(
            "div > span.mr-3.d-none.d-sm-block > a"
        ).text
        publish_time = main_content_element.select_one(
            "div > span:nth-child(2) > span"
        )["title"]
        view = main_content_element.select_one("div > span.views.mr-3").text
        comment = main_content_element.select_one("div > span:nth-child(6) > a").text
        like = main_content_element.select_one(
            "div > span:nth-child(7) > a > span"
        ).text
        print(title, category, publish_time, view, comment, like)

        # 写入Excel
        write_excel(
            title,
            title_url,
            category,
            date_convert(publish_time),
            convert_str_to_num(view),
            convert_str_to_num(comment),
            convert_str_to_num(like),
        )
        time.sleep(1)


if __name__ == "__main__":
    home_url = "https://ai-bot.cn/blog/"
    get_blog_list(home_url)
    for i in range(2, 22):
        url = f"{home_url}page/{i}"
        print(url)
        get_blog_list(url)
